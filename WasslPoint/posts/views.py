import json
from sqlite3 import IntegrityError
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from .models import TrainingOpportunity, Application, Message
from profiles.models import CompanyProfile, StudentProfile, Major, City
from subscriptions.models import has_active_subscription
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.contrib.messages import get_messages
from django.views.decorators.http import require_POST, require_http_methods
from django.urls import reverse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from .forms import TrainingOpportunityForm

# --- Helper Decorators ---
def company_required(view_func):
    @login_required
    def _wrapped_view(request, *args, **kwargs):
        if not hasattr(request.user, 'company_profile') or not request.user.company_profile:
            # Translated Message
            messages.error(request, "تم رفض الوصول. يتوجب عليك الدخول بحساب شركة.")
            return redirect('main:home_view')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def student_required(view_func):
    @login_required
    def _wrapped_view(request, *args, **kwargs):
        if not hasattr(request.user, 'student_profile') or not request.user.student_profile:
            # Translated Message
            messages.error(request, "تم رفض الوصول. يتوجب عليك الدخول بحساب طالب.")
            if hasattr(request.user, 'company_profile') and request.user.company_profile:
                 return redirect('profiles:company_profile_view')
            return redirect('main:home_view')
        return view_func(request, *args, **kwargs)
    return _wrapped_view


# --- Public/Student Views ---

def opportunity_list(request):
    """ Displays active opportunities viewable by anyone. """
    opportunities = TrainingOpportunity.objects.filter(
        status=TrainingOpportunity.Status.ACTIVE,
        application_deadline__gte=timezone.now().date()
    ).select_related('company', 'city').prefetch_related('majors_needed').order_by('-created_at')
    return render(request, 'posts/opportunity_list.html', {'opportunities': opportunities})

def opportunity_detail(request, opportunity_id):
    """
    تفاصيل فرصة تدريبية. متاحة للجميع، مع منطق كامل لزر التقديم
    وظهور الرسائل بحسب حالة الطالب/الاشتراك/الموعد.
    """
    opportunity = get_object_or_404(
        TrainingOpportunity.objects
            .select_related("company", "city")
            .prefetch_related("majors_needed"),
        pk=opportunity_id
    )

    # 0) متغيّرات أساسية
    today              = timezone.now().date()
    application_closed = (
        opportunity.status != TrainingOpportunity.Status.ACTIVE
        or opportunity.application_deadline < today
    )

    # أعلام (flags) يُطلبها القالب
    is_student            = False
    is_opportunity_owner  = False
    already_applied       = False
    can_withdraw          = False
    show_apply_form       = False
    needs_subscription    = False
    show_reapply_form     = False
    application           = None

    # 1) مستخدم مسجَّل؟
    if request.user.is_authenticated:

        # 1‑A) هل هو صاحب الشركة التي أنشأت الفرصة؟
        if hasattr(request.user, "company_profile") and request.user.company_profile:
            is_opportunity_owner = (request.user.company_profile == opportunity.company)

        # 1‑B) هل هو طالب؟
        if hasattr(request.user, "student_profile") and request.user.student_profile:
            is_student = True
            student_profile = request.user.student_profile

            # هل قدّم من قبل؟
            try:
                application = Application.objects.get(
                    opportunity=opportunity,
                    student=student_profile
                )
                already_applied = True

                # هل يحقّ له سحب الطلب؟
                if application.status in (
                    Application.ApplicationStatus.PENDING,
                    Application.ApplicationStatus.ACCEPTED,
                ):
                    can_withdraw = True

                # يمكنه إعادة التقديم إذا سحب الطلب سابقاً + الموعد مفتوح + اشتراك فعّال
                if (
                    application.status == Application.ApplicationStatus.WITHDRAWN
                    and not application_closed
                    and has_active_subscription(request.user)
                ):
                    show_reapply_form = True

            except Application.DoesNotExist:
                # لم يقدِّم بعد
                if not application_closed:
                    if has_active_subscription(request.user):
                        show_apply_form = True
                    else:
                        needs_subscription = True

    # 2) تحضير الـ context للقالب
    context = {
        "opportunity":            opportunity,
        "application_closed":     application_closed,

        "is_student":             is_student,
        "is_opportunity_owner":   is_opportunity_owner,

        "application":            application,
        "already_applied":        already_applied,

        "can_withdraw":           can_withdraw,
        "show_apply_form":        show_apply_form,
        "show_reapply_form":      show_reapply_form,
        "needs_subscription":     needs_subscription,
    }

    return render(request, "posts/opportunity_detail.html", context)

@login_required
@student_required
@require_POST
def apply_opportunity(request, opportunity_id):
    opportunity = get_object_or_404(TrainingOpportunity, pk=opportunity_id)
    student_profile = request.user.student_profile
    application_message = request.POST.get('message', '')

    # 1. Check Subscription
    if not has_active_subscription(request.user):
        # Translated Message
        messages.error(request, "تحتاج إلى اشتراك نشط للتقدم بطلب للحصول على الفرص.")
        return redirect('subscriptions:plans')

    # 2. Check Opportunity Status and Deadline
    if opportunity.status != TrainingOpportunity.Status.ACTIVE:
        # Translated Message
        messages.error(request, "هذه الفرصة لم تعد نشطة.")
        return redirect('posts:opportunity_detail', opportunity_id=opportunity_id)
    if opportunity.application_deadline < timezone.now().date():
        # Translated Message
        messages.error(request, "انتهى الموعد النهائي للتقديم لهذه الفرصة.")
        return redirect('posts:opportunity_detail', opportunity_id=opportunity_id)

    # 3. Check if already applied (or withdrawn)
    defaults = {
        'status': Application.ApplicationStatus.PENDING,
        'message': application_message,
        # 'student_has_seen_latest_status': True, # Assumes student knows initial status
        # 'company_has_seen_application': False, # Assumes company needs notification
    }
    existing_application, created = Application.objects.get_or_create(
        opportunity=opportunity,
        student=student_profile,
        defaults=defaults
    )

    if not created:
        if existing_application.status == Application.ApplicationStatus.WITHDRAWN:
             existing_application.status = Application.ApplicationStatus.PENDING
             existing_application.applied_at = timezone.now()
             existing_application.message = application_message # Update message
             # existing_application.student_has_seen_latest_status = True
             # existing_application.company_has_seen_application = False
             existing_application.save()
             # Translated Message
             messages.success(request, "لقد تمت إعادة التقديم على هذه الفرصة.")
             return redirect('posts:my_applications')
        else:
            # Translated Message
            messages.warning(request, "لقد قمت بالتقديم مسبقاً على هذه الفرصة.")
            return redirect('posts:opportunity_detail', opportunity_id=opportunity_id)
    else:
         # Application was newly created
         # Translated Message
         messages.success(request, "تم إرسال طلبك بنجاح!")
         return redirect('posts:my_applications')

@login_required
@student_required
def my_applications_list(request):
    """ Shows the logged-in student their applications with pagination. """
    student_profile = request.user.student_profile
    application_list = Application.objects.filter(
        student=student_profile
    ).select_related('opportunity__company', 'opportunity__city').order_by('-applied_at')

    # --- Placeholder: Mark unseen status updates as seen ---
    # Application.objects.filter(
    #     student=student_profile,
    #     student_has_seen_latest_status=False # Requires model field
    # ).update(student_has_seen_latest_status=True)
    # --- End Placeholder ---

    paginator = Paginator(application_list, 10)
    page_number = request.GET.get('page')

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)

    context = {
        'page_obj': page_obj
    }
    return render(request, 'posts/my_applications.html', context)


@login_required
@student_required
@require_POST
def withdraw_application(request, application_id):
    application = get_object_or_404(Application, pk=application_id, student=request.user.student_profile)

    if application.status not in [Application.ApplicationStatus.PENDING, Application.ApplicationStatus.ACCEPTED]:
         # Translated Message
         messages.warning(request, f"لا يمكن سحب الطلب وهو بالحالة '{application.get_status_display()}'.")
         return redirect('posts:my_applications')

    application.status = Application.ApplicationStatus.WITHDRAWN
    # application.student_has_seen_latest_status = True # Student initiated, they have seen
    # application.company_has_seen_application = False # Company needs notification?
    application.save()
    # Translated Message
    messages.success(request, "تم سحب الطلب بنجاح.")
    return redirect('posts:my_applications')


# --- Company Views ---

@company_required
def company_dashboard(request):
    """ Displays opportunities created by the logged-in company. """
    company_profile = CompanyProfile.objects.get(user=request.user)
    opportunities = TrainingOpportunity.objects.filter(company=company_profile).prefetch_related('applications').order_by('-created_at')
    context = {
        'opportunities': opportunities,
        'company_profile': company_profile
    }
    return render(request, 'posts/company_dashboard.html', context)

@company_required
def create_opportunity(request):
    if request.method == 'POST':
        is_draft = 'save_draft' in request.POST
        form = TrainingOpportunityForm(request.POST, skip_required=is_draft)
        if form.is_valid():
            opportunity = form.save(commit=False)
            opportunity.status = 'DRAFT' if is_draft else 'ACTIVE'
            opportunity.company = CompanyProfile.objects.get(user=request.user)

            # Skip saving if it's a draft and required fields are missing
            required_fields = ['start_date', 'duration', 'application_deadline']
            missing_required = any(not getattr(opportunity, field) for field in required_fields)

            if is_draft and missing_required:
                # skip saving incomplete drafts
                messages.success(request, 'تم حفظ الفرصة كمسودة')
            else:
                try:
                    opportunity.save()
                    form.save_m2m()
                    if is_draft:
                        messages.success(request, 'تم حفظ الفرصة كمسودة')
                except IntegrityError as e:
                    raise e  # Show error for unexpected cases

            return redirect('posts:company_dashboard')
    else:
        form = TrainingOpportunityForm()

    return render(request, 'posts/create_opportunity.html', {'form': form})


@login_required
@company_required # Ensures only logged-in companies can access
def edit_opportunity(request, opportunity_id):
    opportunity = get_object_or_404(TrainingOpportunity, id=opportunity_id)
    company_profile = request.user.company_profile

    # Authorization: Check if the logged-in company is the owner of the opportunity
    if opportunity.company != company_profile:
        messages.error(request, "لا تملك الصلاحية لتعديل هذه الفرصة.")
        return redirect('posts:company_dashboard') # Or another appropriate redirect

    if request.method == 'POST':
        form = TrainingOpportunityForm(request.POST, instance=opportunity)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تحديث الفرصة بنجاح.')
            return redirect('posts:opportunity_detail', opportunity_id=opportunity.id)
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء الموجودة في النموذج.')
    else: # GET request
        form = TrainingOpportunityForm(instance=opportunity)

    return render(request, 'posts/edit_opportunity.html', {
        'form': form,
        'opportunity': opportunity, # Pass opportunity for the cancel button or other context
        'is_edit': True # You can use this in the template if needed for conditional logic
    })


@require_POST
@login_required
def delete_opportunity(request, opportunity_id):
    opportunity = get_object_or_404(TrainingOpportunity, pk=opportunity_id)

    is_owner = hasattr(request.user, 'company_profile') and opportunity.company == request.user.company_profile
    if not request.user.is_staff and not is_owner:
        # Translated Message
        messages.error(request, "لا تملك الصلاحية لحذف هذه الفرصة.")
        return redirect(request.META.get('HTTP_REFERER', reverse('posts:opportunity_list')))

    try:
        opportunity_name = str(opportunity)
        opportunity.delete()
        # Translated Message
        messages.success(request, f"تم حذف الفرصة '{opportunity_name}' بنجاح.")
    except Exception as e:
        # Translated Message
        messages.error(request, f"حدث خطأ أثناء حذف الفرصة: {e}")
        return redirect(request.META.get('HTTP_REFERER', reverse('posts:opportunity_detail', args=[opportunity_id])))

    if request.user.is_staff and not is_owner:
         return redirect('posts:opportunity_list')
    else:
        return redirect('posts:company_dashboard')



@login_required
def opportunity_applications(request, opportunity_id):
    """ Company or Admin views applications for a specific opportunity. """
    opportunity = get_object_or_404(TrainingOpportunity, pk=opportunity_id)

    # --- Authorization Check (THIS IS CORRECT) ---
    is_owner = hasattr(request.user, 'company_profile') and opportunity.company == request.user.company_profile
    if not request.user.is_staff and not is_owner:
        messages.error(request, "ليس لديك الإذن لعرض المتقدمين على هذه الفرصة.")
        # Consider redirecting instead of raising PermissionDenied if you want a friendlier message flow
        # return redirect('main:home_view') # Or company dashboard
        raise PermissionDenied("ليس لديك الإذن لعرض المتقدمين على هذه الفرصة.")
    # --- End Authorization Check ---

    applications_qs = Application.objects.filter(
        opportunity=opportunity
    ).select_related(
        'student__user',
        'student__personal_info', # For Name
        'student__contact_info'   # For Phone
        ).order_by('status', '-applied_at')

    context = {
        'opportunity': opportunity,
        'applications': applications_qs,
        'statuses': Application.ApplicationStatus.choices
    }
    return render(request, 'posts/opportunity_applications.html', context)


@require_POST
@login_required
def update_application_status(request, application_id):
    """ Company or Admin updates the status of an application. """
    application = get_object_or_404(Application.objects.select_related('opportunity', 'student'), pk=application_id)
    opportunity = application.opportunity

    is_owner = hasattr(request.user, 'company_profile') and opportunity.company == request.user.company_profile
    if not request.user.is_staff and not is_owner:
        # Translated Message
        messages.error(request, "لا تملك الصلاحية للتحديث على حالة التقديم.")
        return redirect('posts:opportunity_applications', opportunity_id=opportunity.id)

    new_status = request.POST.get('status')
    current_status = application.status

    if new_status in Application.ApplicationStatus.values:
        if application.status == Application.ApplicationStatus.WITHDRAWN and request.user == application.student.user:
             # Translated Message
             messages.error(request, "لا يمكن تحديث الحالة من 'مسحوب' هنا.")
        else:
            application.status = new_status
            # --- Placeholder: Mark as unseen by student ---
            # if new_status != current_status:
            #      application.student_has_seen_latest_status = False # Requires model field
            # --- End Placeholder ---
            application.save()
            # Translated Message
            messages.success(request, f"تم تحديث حالة الطلب إلى '{application.get_status_display()}'.")
            # TODO: Notify student
    else:
        # Translated Message
        messages.error(request, "تم اختيار حالة غير صالحة.")

    return redirect('posts:opportunity_applications', opportunity_id=opportunity.id)


# --- Chat View ---
@login_required
@login_required
def application_chat(request, application_id):
    """ Displays chat, marks messages read, handles sending messages.
        Alert filtering is now handled in base.html.
    """
    # REMOVED the message filtering block from here

    application = get_object_or_404(
        Application.objects.select_related("student__user", "opportunity__company__user"),
        pk=application_id,
    )

    # Permission Check
    is_student = hasattr(request.user, 'student_profile') and request.user == application.student.user
    is_company = hasattr(request.user, 'company_profile') and request.user == application.opportunity.company.user
    is_admin = request.user.is_staff

    if not (is_student or is_company or is_admin):
        messages.error(request,"لا تملك الصلاحية للوصول الى هذه المحادثة.")
        raise PermissionDenied("لا تملك الصلاحية للوصول الى هذه المحادثة.")

    # Fetch actual chat messages from the database for display
    messages_qs = Message.objects.filter(application=application).select_related("sender").order_by("sent_at")

    # Mark messages in the database as read by the current user
    if is_student or is_company:
        recipient_filter = Q(application=application, is_read=False) & ~Q(sender=request.user)
        Message.objects.filter(recipient_filter).update(is_read=True)

    # Handle sending a new message (POST request part)
    if request.method == "POST":
        content = request.POST.get("content", "").strip()
        if content:
            Message.objects.create(
                application=application, sender=request.user, content=content
            )
            # OPTIONAL: Add a success message if desired
            # messages.success(request, "تم إرسال رسالتك بنجاح!")
            return redirect("posts:application_chat", application_id=application_id)
        else:
            messages.warning(request, "لا يمكن إرسال رسالة فارغة!")
            return redirect("posts:application_chat", application_id=application_id)

    # Prepare context for rendering the template (GET request part)
    context = {
        "application": application,
        "messages_qs": messages_qs, # Pass the actual chat messages
        "is_student_view": is_student,
        "is_company_view": is_company,
        # No need to pass 'alert_messages' separately anymore
    }
    return render(request, "posts/application_chat.html", context)

@login_required
# posts/views.py

@login_required
def export_opportunity_applications_excel(request, opportunity_id):
    opportunity = get_object_or_404(TrainingOpportunity, pk=opportunity_id)
    user = request.user

    is_owner = hasattr(user, 'company_profile') and opportunity.company == user.company_profile
    if not user.is_staff and not is_owner:
        messages.error(request, "ليس لديك الصلاحية لتصدير قائمة المتقدمين لهذه الفرصة.")
        return redirect('posts:opportunity_applications', opportunity_id=opportunity_id)

    applications = Application.objects.filter(
        opportunity=opportunity
    ).select_related(
        'student__user',
        'student__personal_info',
        'student__contact_info'
    ).order_by('status', '-applied_at')

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Make sheet title more descriptive
    sheet.title = f"Applicants for {opportunity.title[:20]}" # Sheet titles have length limits
    sheet.sheet_view.rightToLeft = True

    # --- Add Opportunity Title as a Header in the Sheet ---
    # Merge cells for the title
    sheet.merge_cells('A1:F1') # Merge across the width of your headers
    title_cell = sheet['A1']
    title_cell.value = f"قائمة المتقدمين لفرصة: {opportunity.title}"
    title_cell.font = Font(bold=True, size=14, name='Arial') # Example styling
    title_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 20 # Adjust row height for the title

    # Define Headers (starting from row 2 now)
    headers = [
        "اسم المتقدم", "البريد الإلكتروني", "رقم الهاتف",
        "تاريخ التقديم", "الحالة", "رسالة التقديم"
    ]
    sheet.append(headers) # This will append to the next available row, which is row 2

    # Style Header Row for applicant data (Row 2)
    header_font = Font(bold=True, name='Arial') # Consistent font
    for col_num, header_text in enumerate(headers, 1): # header_text instead of header
        cell = sheet.cell(row=2, column=col_num) # Explicitly set to row 2
        cell.font = header_font
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')


    # Populate Data (starting from row 3)
    current_row = 3 # Start data from row 3
    for app in applications:
        student_name = getattr(app.student.personal_info, 'full_name', None) or \
                       app.student.user.get_full_name() or \
                       app.student.user.username
        email = getattr(app.student.user, 'email', 'N/A')
        phone = getattr(getattr(app.student, 'contact_info', None), 'phone', 'N/A')
        applied_date = app.applied_at.strftime('%Y-%m-%d %H:%M') if app.applied_at else 'N/A'
        status_display = app.get_status_display() # Renamed for clarity
        message_text = getattr(app, 'message', '') # Renamed for clarity

        sheet.cell(row=current_row, column=1, value=student_name)
        sheet.cell(row=current_row, column=2, value=email)
        sheet.cell(row=current_row, column=3, value=phone)
        sheet.cell(row=current_row, column=4, value=applied_date)
        sheet.cell(row=current_row, column=5, value=status_display)
        sheet.cell(row=current_row, column=6, value=message_text)
        current_row += 1


    # Adjust Column Widths (Optional)
    for col_num in range(1, sheet.max_column + 1):
         column_letter = get_column_letter(col_num)
         max_length = 0
         # Iterate from row 2 (headers) downwards for calculating max_length
         for row_idx in range(2, sheet.max_row + 1):
             cell_value = sheet.cell(row=row_idx, column=col_num).value
             if cell_value:
                 cell_len = len(str(cell_value))
                 if cell_len > max_length:
                     max_length = cell_len
         adjusted_width = (max_length + 4) # Add some padding
         if adjusted_width > 50: # Optional: cap max width
             adjusted_width = 50
         sheet.column_dimensions[column_letter].width = adjusted_width


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_company_name = "".join(c if c.isalnum() else "_" for c in opportunity.company.company_name)
    # Make filename more descriptive and include opportunity title if possible (sanitize it)
    safe_opportunity_title = "".join(c if c.isalnum() else "_" for c in opportunity.title[:30]) # Truncate and sanitize
    filename = f'متقدمون_{safe_opportunity_title}_{safe_company_name}_{opportunity_id}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    workbook.save(response)
    return response